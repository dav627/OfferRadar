#!/usr/bin/env python3
"""
秋招Agent 统一启动器

用法:
  python3 launcher.py run              执行一次完整流程
  python3 launcher.py run --lite       轻量模式（无需Playwright）
  python3 launcher.py run --no-gmail   跳过Gmail检查
  python3 launcher.py run --no-push    跳过推送

  python3 launcher.py init             首次初始化（生成Excel等）
  python3 launcher.py status           查看系统状态
  python3 launcher.py test-push        发送测试推送
  python3 launcher.py test-llm         测试 LLM 接口
  python3 launcher.py gmail-auth       Gmail OAuth 授权

  python3 launcher.py schedule         查看定时任务状态
  python3 launcher.py schedule --on    开启定时任务（每天执行）
  python3 launcher.py schedule --off   关闭定时任务
  python3 launcher.py schedule --time 08:30  设置执行时间
"""

import sys
import os
import json
import subprocess
import asyncio
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent
sys.path.insert(0, str(BASE_DIR))


def cmd_run():
    """执行一次完整的抓取+播报+推送流程"""
    import run_daily
    asyncio.run(run_daily.main())


def cmd_init():
    """首次初始化：生成Excel表、检查依赖"""
    print("=== 秋招Agent 初始化 ===\n")

    # Check dependencies
    print("[1/3] 检查依赖...")
    deps = {"openpyxl": False, "playwright": False}
    for pkg in deps:
        try:
            __import__(pkg)
            deps[pkg] = True
            print(f"  [OK] {pkg}")
        except ImportError:
            print(f"  [MISS] {pkg}")

    if not deps["openpyxl"]:
        print("\n  必需依赖缺失，请运行: pip3 install openpyxl")
        return

    # Check .env
    print("\n[2/3] 检查配置...")
    env_path = BASE_DIR / ".env"
    if env_path.exists():
        print("  [OK] .env 已配置")
    else:
        print("  [INFO] 未找到 .env，已从模板创建")
        import shutil
        shutil.copy(BASE_DIR / ".env.example", env_path)

    # Generate Excel
    print("\n[3/3] 生成投递跟踪表...")
    excel_path = BASE_DIR / "秋招投递跟踪表.xlsx"
    if excel_path.exists():
        print(f"  [SKIP] {excel_path.name} 已存在")
    else:
        _generate_excel()
        print(f"  [OK] 已生成 {excel_path.name}")

    # Create runtime dirs
    for d in ["抓取结果", "每日播报", "gmail_results", "cookies", "logs"]:
        (BASE_DIR / d).mkdir(exist_ok=True)

    print("\n=== 初始化完成 ===")
    print("下一步: python3 launcher.py run --lite --no-gmail --no-push")


def _generate_excel():
    """生成空的投递跟踪表"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投递跟踪表"

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['公司', '梯队', '类别', '部门/团队', '岗位名称', '工作地点',
               'JD核心要求', '匹配度', '投递链接', '内推信息', '网申开放',
               '网申截止', '薪资范围', '投递状态', '笔试时间', '面试进度', '备注']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    col_widths = [12, 6, 14, 16, 22, 10, 30, 8, 35, 20, 12, 12, 12, 10, 12, 15, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Load companies
    with open(BASE_DIR / "公司清单.json") as f:
        data = json.load(f)

    for row_idx, c in enumerate(data["companies"], 2):
        ws.cell(row=row_idx, column=1, value=c["name"])
        ws.cell(row=row_idx, column=2, value=c.get("tier", ""))
        ws.cell(row=row_idx, column=3, value=c.get("category", ""))
        ws.cell(row=row_idx, column=14, value="待关注")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data['companies'])+1}"

    wb.save(BASE_DIR / "秋招投递跟踪表.xlsx")


def cmd_status():
    """查看系统状态"""
    print("=== 秋招Agent 系统状态 ===\n")

    # Config
    env_ok = (BASE_DIR / ".env").exists()
    cred_ok = (BASE_DIR / "credentials.json").exists()
    token_ok = (BASE_DIR / "token.json").exists()
    excel_ok = (BASE_DIR / "秋招投递跟踪表.xlsx").exists()

    print(f"  .env 配置文件:     {'已配置' if env_ok else '未配置 (复制 .env.example 并填写)'}")
    print(f"  Gmail credentials: {'已配置' if cred_ok else '未配置'}")
    print(f"  Gmail token:       {'已授权' if token_ok else '未授权 (运行 launcher.py gmail-auth)'}")
    print(f"  投递跟踪表:        {'已生成' if excel_ok else '未生成 (运行 launcher.py init)'}")

    # Push channels
    import config_loader
    push = config_loader.get_push_config()
    active = [k for k, v in push.items() if v.get("enabled")]
    print(f"  推送渠道:          {', '.join(active) if active else '未配置'}")

    # LLM
    llm = config_loader.get_llm_config()
    if llm["api_key"]:
        print(f"  LLM 接口:          {llm['model']} @ {llm['base_url'][:40]}")
    else:
        print(f"  LLM 接口:          未配置 (在 config.yaml 的 llm.api_key 填写)")

    # Proxy
    proxy = config_loader.get_proxy()
    print(f"  代理设置:          {proxy.get('http', '未设置')}")

    # Schedule
    sched = config_loader.get_schedule_config()
    print(f"  定时任务:          {'已启用 ' + sched['time'] if sched['enabled'] else '未启用'}")

    # Recent data
    latest = BASE_DIR / "抓取结果" / "latest.json"
    if latest.exists():
        with open(latest) as f:
            d = json.load(f)
        print(f"\n  最近抓取: {d.get('scraped_at', 'N/A')[:16]} | {d.get('total_jobs', 0)} 条岗位")

    reports = sorted((BASE_DIR / "每日播报").glob("*.md")) if (BASE_DIR / "每日播报").exists() else []
    if reports:
        print(f"  最近播报: {reports[-1].stem}")

    # Companies
    with open(BASE_DIR / "公司清单.json") as f:
        companies = json.load(f)["companies"]
    print(f"\n  监控公司数: {len(companies)}")
    print(f"  目标方向:   LLM应用算法（排除多模态/基模）")
    print(f"  招聘类型:   仅校招（2027届）")


def cmd_test_push():
    """发送测试推送"""
    from notifier import send_notification
    send_notification("秋招Agent测试", "这是一条测试消息，收到说明推送配置成功！\n\n时间: " + datetime.now().strftime("%Y-%m-%d %H:%M"))


def cmd_gmail_auth():
    """Gmail OAuth 授权"""
    if not (BASE_DIR / "credentials.json").exists():
        print("[ERROR] 未找到 credentials.json")
        print("请参考 README.md 中 Gmail 配置步骤获取")
        return
    subprocess.run([sys.executable, str(BASE_DIR / "gmail_auth.py")])


def cmd_schedule(args):
    """定时任务管理"""
    import config_loader

    sched = config_loader.get_schedule_config()

    if "--on" in args:
        sched["enabled"] = True
        config_loader.save_schedule_config(sched)
        h, m = sched["time"].split(":")
        subprocess.run(["bash", str(BASE_DIR / "setup_schedule.sh"), h, m])
        print(f"[OK] 定时任务已开启，每天 {sched['time']} 执行")

    elif "--off" in args:
        sched["enabled"] = False
        config_loader.save_schedule_config(sched)
        plist_name = f"com.{os.environ.get('USER', 'user')}.qiuzhao-agent"
        plist_path = Path.home() / "Library" / "LaunchAgents" / f"{plist_name}.plist"
        subprocess.run(["launchctl", "unload", str(plist_path)], capture_output=True)
        print("[OK] 定时任务已关闭")

    elif "--time" in args:
        idx = args.index("--time")
        if idx + 1 < len(args):
            sched["time"] = args[idx + 1]
            config_loader.save_schedule_config(sched)
            print(f"[OK] 执行时间已设为 {sched['time']}")
            if sched["enabled"]:
                h, m = sched["time"].split(":")
                subprocess.run(["bash", str(BASE_DIR / "setup_schedule.sh"), h, m])

    else:
        print(f"定时任务: {'已启用' if sched['enabled'] else '未启用'}")
        print(f"执行时间: {sched['time']}")
        print(f"运行模式: {sched['mode']}")
        print(f"Gmail:    {'开' if sched['gmail'] else '关'}")
        print(f"推送:     {'开' if sched['push'] else '关'}")
        print()
        print("管理命令:")
        print("  launcher.py schedule --on         开启")
        print("  launcher.py schedule --off        关闭")
        print("  launcher.py schedule --time 08:30 改时间")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return

    cmd = sys.argv[1]
    remaining = sys.argv[2:]

    # Pass flags through to run_daily
    if cmd == "run":
        sys.argv = ["run_daily.py"] + remaining
        import run_daily
        run_daily.USE_LITE = "--lite" in remaining
        run_daily.SKIP_GMAIL = "--no-gmail" in remaining
        run_daily.SKIP_PUSH = "--no-push" in remaining
        asyncio.run(run_daily.main())
    elif cmd == "init":
        cmd_init()
    elif cmd == "status":
        cmd_status()
    elif cmd == "test-push":
        cmd_test_push()
    elif cmd == "test-llm":
        from llm_analyzer import check_llm_available
        print("[INFO] 测试 LLM 连接...")
        if check_llm_available():
            print("[OK] LLM API 可用")
        else:
            llm = config_loader.get_llm_config()
            if not llm["api_key"]:
                print("[FAIL] 未配置 LLM_API_KEY，请在 .env 中设置")
            else:
                print(f"[FAIL] 无法连接 {llm['base_url']}")
    elif cmd == "gmail-auth":
        cmd_gmail_auth()
    elif cmd == "schedule":
        cmd_schedule(remaining)
    else:
        print(f"未知命令: {cmd}")
        print(__doc__)


if __name__ == "__main__":
    main()

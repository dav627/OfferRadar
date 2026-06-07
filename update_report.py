#!/usr/bin/env python3
"""
每日播报生成 + Excel更新脚本
对比前后抓取数据，生成增量报告，更新投递跟踪表
"""

import json
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = Path(__file__).parent
RESULTS_DIR = BASE_DIR / "抓取结果"
REPORT_DIR = BASE_DIR / "每日播报"
EXCEL_PATH = BASE_DIR / "秋招投递跟踪表.xlsx"
REPORT_DIR.mkdir(exist_ok=True)


def load_latest_results() -> Optional[dict]:
    latest = RESULTS_DIR / "latest.json"
    if latest.exists():
        with open(latest, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def load_previous_results() -> Optional[dict]:
    """Load the second-most-recent results for comparison"""
    files = sorted(RESULTS_DIR.glob("scrape_*.json"), reverse=True)
    if len(files) >= 2:
        with open(files[1], "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def find_new_jobs(latest: dict, previous: Optional[dict]) -> list:
    """Find jobs in latest that weren't in previous"""
    if not previous:
        return latest.get("jobs", [])

    prev_titles = {(j["company"], j["title"]) for j in previous.get("jobs", [])}
    new_jobs = []
    for job in latest.get("jobs", []):
        if (job["company"], job["title"]) not in prev_titles:
            new_jobs.append(job)
    return new_jobs


def generate_daily_report(latest: dict, new_jobs: list) -> str:
    """Generate markdown daily report"""
    today = datetime.now().strftime("%Y-%m-%d")
    all_jobs = latest.get("jobs", [])

    # Group by company
    by_company = {}
    for job in all_jobs:
        company = job["company"]
        if company not in by_company:
            by_company[company] = []
        by_company[company].append(job)

    new_by_company = {}
    for job in new_jobs:
        company = job["company"]
        if company not in new_by_company:
            new_by_company[company] = []
        new_by_company[company].append(job)

    report = f"""# 秋招每日播报 - {today}

## 今日概况

| 指标 | 数值 |
|------|------|
| 抓取时间 | {latest.get('scraped_at', 'N/A')} |
| 抓取岗位总数 | {len(all_jobs)} |
| 新增岗位数 | {len(new_jobs)} |
| 涉及公司数 | {len(by_company)} |

---

"""

    if new_jobs:
        report += "## 新增岗位\n\n"
        report += "| 公司 | 岗位 | 地点 | 来源 |\n"
        report += "|------|------|------|------|\n"
        for job in new_jobs:
            report += f"| {job['company']} | {job['title']} | {job.get('location', '-')} | {job.get('source', '-')} |\n"
        report += "\n---\n\n"

    if new_by_company:
        report += "## 新增详情\n\n"
        for company, jobs in new_by_company.items():
            report += f"### {company} (+{len(jobs)})\n\n"
            for job in jobs:
                report += f"- **{job['title']}**"
                if job.get("location"):
                    report += f" | {job['location']}"
                if job.get("department"):
                    report += f" | {job['department']}"
                report += "\n"
            report += "\n"

    # Summary of all active companies
    report += "## 当前所有在招公司\n\n"
    report += "| 公司 | 在招岗位数 |\n"
    report += "|------|------------|\n"
    for company in sorted(by_company.keys(), key=lambda x: len(by_company[x]), reverse=True):
        report += f"| {company} | {len(by_company[company])} |\n"

    # LLM smart summary (if configured)
    try:
        from llm_analyzer import generate_smart_report, analyze_jobs
        if new_jobs:
            new_jobs = analyze_jobs(new_jobs)
        smart = generate_smart_report(all_jobs, new_jobs)
        if smart:
            report += "\n## AI 分析\n\n" + smart + "\n"
    except Exception as e:
        print(f"[INFO] LLM 分析跳过: {e}")

    report += f"\n---\n\n*生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}*\n"

    return report


def update_excel(latest: dict):
    """Update Excel with new findings"""
    if not EXCEL_PATH.exists():
        print("[WARN] Excel文件不存在，跳过更新")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["投递跟踪表"]

    # Read existing companies in Excel
    existing = {}
    for row in range(2, ws.max_row + 1):
        company = ws.cell(row=row, column=1).value
        if company:
            existing[company] = row

    jobs = latest.get("jobs", [])
    by_company = {}
    for job in jobs:
        c = job["company"]
        if c not in by_company:
            by_company[c] = []
        by_company[c].append(job)

    # Update status for companies where we found active jobs
    for company, jobs in by_company.items():
        if company in existing:
            row = existing[company]
            current_status = ws.cell(row=row, column=14).value
            if current_status == "待关注":
                ws.cell(row=row, column=14, value="待投递")
                ws.cell(row=row, column=14).font = Font(bold=True, color="FF8C00")
                # Update notes
                current_note = ws.cell(row=row, column=17).value or ""
                new_note = f"[{datetime.now().strftime('%m-%d')}] 发现{len(jobs)}个相关岗位"
                ws.cell(row=row, column=17, value=f"{new_note}; {current_note}" if current_note else new_note)

    # Update summary sheet timestamp
    if "汇总统计" in wb.sheetnames:
        ws2 = wb["汇总统计"]
        ws2['A3'] = f'更新时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

    wb.save(EXCEL_PATH)
    print(f"[DONE] Excel已更新: {EXCEL_PATH}")


def run():
    print(f"[INFO] 开始生成每日播报 - {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    latest = load_latest_results()
    if not latest:
        print("[ERROR] 未找到抓取结果，请先运行 scraper.py")
        return

    previous = load_previous_results()
    new_jobs = find_new_jobs(latest, previous)

    # Generate report
    report = generate_daily_report(latest, new_jobs)
    today = datetime.now().strftime("%Y-%m-%d")
    report_path = REPORT_DIR / f"{today}.md"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"[DONE] 播报已生成: {report_path}")

    # Update Excel
    update_excel(latest)

    return report


if __name__ == "__main__":
    run()

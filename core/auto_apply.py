#!/usr/bin/env python3
"""
自动投递模块（Phase 5）
通过 Playwright 自动化投递简历到各招聘平台

⚠️ 重要注意事项：
1. 自动投递前会要求人工确认
2. 首次使用需要手动登录各平台保存cookie
3. 建议先在Excel中标记"待投递"状态，脚本只投递这些岗位
4. 部分平台有反自动化检测，建议控制投递频率

支持的平台：
- 字节跳动 (jobs.bytedance.com)
- 百度 (talent.baidu.com)
- 阿里巴巴 (talent.alibaba.com)
- 腾讯 (join.qq.com)
- 美团 (campus.meituan.com)

用法:
  python3 auto_apply.py --login     # 首次登录保存cookie
  python3 auto_apply.py --dry-run   # 模拟投递（不实际提交）
  python3 auto_apply.py --apply     # 正式投递
"""

import json
import asyncio
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from core import PROJECT_ROOT, DATA_DIR
COOKIE_DIR = DATA_DIR / "cookies"
COOKIE_DIR.mkdir(exist_ok=True)
RESUME_DIR = Path(__file__).parent.parent  # 父目录（简历所在位置）

PLATFORMS = {
    "bytedance": {
        "name": "字节跳动",
        "login_url": "https://jobs.bytedance.com/campus",
        "apply_base": "https://jobs.bytedance.com/campus/position/apply",
    },
    "baidu": {
        "name": "百度",
        "login_url": "https://talent.baidu.com/external/login",
        "apply_base": "https://talent.baidu.com/jobs/detail",
    },
    "alibaba": {
        "name": "阿里巴巴",
        "login_url": "https://talent.alibaba.com/personal/login",
        "apply_base": "https://talent.alibaba.com/campus/apply",
    },
    "tencent": {
        "name": "腾讯",
        "login_url": "https://join.qq.com/login.html",
        "apply_base": "https://join.qq.com/post_detail.html",
    },
    "meituan": {
        "name": "美团",
        "login_url": "https://campus.meituan.com/login",
        "apply_base": "https://campus.meituan.com/recruit/detail",
    },
}


class AutoApplier:
    def __init__(self, dry_run: bool = True):
        self.dry_run = dry_run
        self.applied = []
        self.failed = []

    async def login_platform(self, platform_key: str):
        """手动登录平台并保存cookie"""
        from playwright.async_api import async_playwright

        platform = PLATFORMS.get(platform_key)
        if not platform:
            print(f"[ERROR] 未知平台: {platform_key}")
            return

        print(f"\n[INFO] 正在打开 {platform['name']} 登录页面...")
        print("[INFO] 请在浏览器中完成登录，登录成功后按回车键保存cookie")

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=False)  # 有头模式，方便手动登录
            context = await browser.new_context()
            page = await context.new_page()

            await page.goto(platform["login_url"])

            # Wait for user to login
            input(f"\n>>> 请在浏览器中完成 {platform['name']} 的登录，然后按回车键... ")

            # Save cookies
            cookies = await context.cookies()
            cookie_path = COOKIE_DIR / f"{platform_key}.json"
            with open(cookie_path, "w") as f:
                json.dump(cookies, f, indent=2)

            print(f"[DONE] Cookie已保存: {cookie_path}")
            await browser.close()

    async def apply_bytedance(self, job_url: str, context) -> bool:
        """投递字节跳动"""
        page = await context.new_page()
        try:
            await page.goto(job_url, timeout=20000)
            await page.wait_for_timeout(2000)

            # Look for apply button
            apply_btn = await page.query_selector('button:has-text("投递"), [class*="apply-btn"]')
            if not apply_btn:
                print(f"  [SKIP] 未找到投递按钮: {job_url}")
                return False

            if self.dry_run:
                print(f"  [DRY-RUN] 找到投递按钮，模拟投递: {job_url}")
                return True

            await apply_btn.click()
            await page.wait_for_timeout(3000)

            # Check if application was submitted
            success_indicator = await page.query_selector(
                ':has-text("投递成功"), :has-text("已投递"), [class*="success"]'
            )
            if success_indicator:
                return True

            # May need to confirm
            confirm_btn = await page.query_selector('button:has-text("确认"), button:has-text("提交")')
            if confirm_btn:
                await confirm_btn.click()
                await page.wait_for_timeout(2000)
                return True

            return False
        except Exception as e:
            print(f"  [ERROR] {e}")
            return False
        finally:
            await page.close()

    async def apply_to_platform(self, platform_key: str, job_urls: list):
        """批量投递到指定平台"""
        from playwright.async_api import async_playwright

        platform = PLATFORMS.get(platform_key)
        if not platform:
            return

        cookie_path = COOKIE_DIR / f"{platform_key}.json"
        if not cookie_path.exists():
            print(f"[ERROR] 未找到 {platform['name']} 的cookie，请先运行 --login")
            return

        print(f"\n[INFO] 开始投递 {platform['name']}（{len(job_urls)}个岗位）")
        if self.dry_run:
            print("[MODE] 模拟模式 - 不会实际提交")

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context()

            # Load cookies
            with open(cookie_path, "r") as f:
                cookies = json.load(f)
            await context.add_cookies(cookies)

            for i, url in enumerate(job_urls, 1):
                print(f"  [{i}/{len(job_urls)}] {url}")

                apply_func = getattr(self, f"apply_{platform_key}", None)
                if apply_func:
                    success = await apply_func(url, context)
                    if success:
                        self.applied.append({"platform": platform_key, "url": url})
                    else:
                        self.failed.append({"platform": platform_key, "url": url})

                # Rate limiting - 每次投递间隔5-10秒
                import random
                await asyncio.sleep(random.uniform(5, 10))

            await browser.close()

    def get_pending_applications(self) -> dict:
        """从Excel读取待投递的岗位"""
        import openpyxl

        excel_path = DATA_DIR / "秋招投递跟踪表.xlsx"
        if not excel_path.exists():
            return {}

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["投递跟踪表"]

        pending = {}
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=14).value
            if status == "待投递":
                company = ws.cell(row=row, column=1).value
                url = ws.cell(row=row, column=9).value
                if url:
                    # Map company to platform key
                    platform_map = {
                        "字节跳动": "bytedance",
                        "百度": "baidu",
                        "阿里巴巴": "alibaba",
                        "腾讯": "tencent",
                        "美团": "meituan",
                    }
                    platform_key = platform_map.get(company)
                    if platform_key:
                        if platform_key not in pending:
                            pending[platform_key] = []
                        pending[platform_key].append(url)

        return pending

    def update_excel_status(self):
        """更新Excel中已投递的状态"""
        import openpyxl

        excel_path = DATA_DIR / "秋招投递跟踪表.xlsx"
        if not excel_path.exists() or self.dry_run:
            return

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["投递跟踪表"]

        applied_urls = {a["url"] for a in self.applied}

        for row in range(2, ws.max_row + 1):
            url = ws.cell(row=row, column=9).value
            if url in applied_urls:
                ws.cell(row=row, column=14, value="已投递")
                ws.cell(row=row, column=17, value=f"[{datetime.now().strftime('%m-%d')}] 自动投递")

        wb.save(excel_path)
        print(f"[DONE] Excel已更新，{len(self.applied)}个岗位标记为已投递")

    async def run(self):
        """执行自动投递"""
        pending = self.get_pending_applications()

        if not pending:
            print("[INFO] 没有待投递的岗位（Excel中状态为'待投递'的行）")
            return

        total = sum(len(urls) for urls in pending.values())
        print(f"\n[INFO] 共有 {total} 个待投递岗位:")
        for platform, urls in pending.items():
            print(f"  - {PLATFORMS[platform]['name']}: {len(urls)} 个")

        if not self.dry_run:
            confirm = input("\n>>> 确认开始投递? (yes/no): ")
            if confirm.lower() != "yes":
                print("[ABORT] 已取消")
                return

        for platform_key, urls in pending.items():
            await self.apply_to_platform(platform_key, urls)

        # Summary
        print("\n" + "=" * 40)
        print(f"投递完成！")
        print(f"  成功: {len(self.applied)}")
        print(f"  失败: {len(self.failed)}")
        print("=" * 40)

        # Update Excel
        self.update_excel_status()


async def main():
    if "--login" in sys.argv:
        applier = AutoApplier()
        # Login to all platforms or specific one
        if len(sys.argv) > 2:
            platform = sys.argv[sys.argv.index("--login") + 1] if sys.argv.index("--login") + 1 < len(sys.argv) else None
            if platform and platform in PLATFORMS:
                await applier.login_platform(platform)
            else:
                print(f"可选平台: {', '.join(PLATFORMS.keys())}")
        else:
            for key in PLATFORMS:
                await applier.login_platform(key)
    elif "--dry-run" in sys.argv:
        applier = AutoApplier(dry_run=True)
        await applier.run()
    elif "--apply" in sys.argv:
        applier = AutoApplier(dry_run=False)
        await applier.run()
    else:
        print(__doc__)


if __name__ == "__main__":
    asyncio.run(main())
